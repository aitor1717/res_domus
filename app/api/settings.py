"""
Settings endpoints: manual monthly budget overrides, Anthropic API key
storage/status, and manual re-import of review/output CSVs into the DB.
"""

import io
import sqlite3
from pathlib import Path
from flask import Blueprint, jsonify, request, current_app, send_file

from csv_safety import sanitize_cell
from db_settings import get_setting, set_setting
from parser.build_db import run_import

bp = Blueprint("settings", __name__, url_prefix="/api/settings")


def _db():
    return sqlite3.connect(current_app.config["DB_PATH"])


@bp.post("/budget")
def set_budget():
    data = request.get_json(force=True)
    month = data.get("month")
    manual_budget = data.get("manual_budget")
    if not month:
        return jsonify({"error": "month required"}), 400

    conn = _db()
    if manual_budget is not None:
        conn.execute("""
            INSERT INTO budget (month, manual_budget)
            VALUES (?, ?)
            ON CONFLICT(month) DO UPDATE SET manual_budget=excluded.manual_budget, updated_ts=CURRENT_TIMESTAMP
        """, (month, float(manual_budget)))
    else:
        conn.execute("DELETE FROM budget WHERE month = ?", (month,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "month": month})


@bp.get("/api-key")
def get_api_key_status():
    db_path = current_app.config["DB_PATH"]
    stored = get_setting(db_path, "anthropic_api_key")
    has_env = bool(current_app.config["ANTHROPIC_API_KEY"])
    return jsonify({"configured": bool(stored or has_env), "source": "settings" if stored else ("env" if has_env else None)})


@bp.post("/api-key")
def set_api_key():
    data = request.get_json(force=True)
    key = (data.get("api_key") or "").strip()
    set_setting(current_app.config["DB_PATH"], "anthropic_api_key", key)
    return jsonify({"ok": True, "configured": bool(key)})


@bp.get("/export")
def export_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    conn = _db()
    cur = conn.execute("""
        SELECT datetime, raw_name, matched_id, matched_category,
               quantity, unit_price, total_price, source
        FROM purchases
        WHERE raw_name != 'TOTAL'
        ORDER BY datetime DESC
    """)
    rows = cur.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "purchases"

    headers = ["date", "raw name", "item (canon)", "category", "qty", "unit price", "total", "source"]
    hdr_fill = PatternFill("solid", fgColor="1E2127")
    hdr_font = Font(name="Calibri", bold=True, color="FF9D6E")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="left")

    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=sanitize_cell(val))

    col_widths = [14, 28, 22, 18, 7, 12, 12, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name="res_domus_purchases.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.post("/reimport")
def reimport():
    db_path = Path(current_app.config["DB_PATH"])
    review_dir = Path(current_app.config["REVIEW_DIR"])
    output_dir = Path(current_app.config["OUTPUT_DIR"])

    r1 = run_import(db_path, review_dir)
    r2 = run_import(db_path, output_dir)
    return jsonify({
        "inserted": r1["inserted"] + r2["inserted"],
        "skipped_dup": r1["skipped_dup"] + r2["skipped_dup"],
        "files_processed": r1["files_processed"] + r2["files_processed"],
        "message": f"review: {r1['message']} | output: {r2['message']}",
    })
